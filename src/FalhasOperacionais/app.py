import re
import shutil
from pathlib import Path
from openpyxl import load_workbook


# -------------------------------
# Diretórios
# -------------------------------
DOWNLOADS = Path.home() / "Downloads"
DESTINO_REDE = Path(r"\\brsbesrv960\Publico\REPORTS\ITAU\FALHAS_OPERACIONAIS")


# -------------------------------
# Regex
# -------------------------------
ARQUIVO_REGEX = re.compile(
    r"^falhasoperacionaiscartões.*\.xlsx$",
    re.IGNORECASE
)

# Ex: 2 a 7.12 | 2.12 a 7.12 | 11.11 a 29.11
SHEET_PERIODO_REGEX = re.compile(
    r"\b\d{1,2}(\.\d{1,2})?\s*a\s*\d{1,2}\.\d{1,2}\b"
)


# -------------------------------
# Funções
# -------------------------------
def tratar_excel(caminho_arquivo: Path):
    """
    Mantém somente a sheet do período,
    renomeia para Plan1 e remove filtros
    """
    wb = load_workbook(
        caminho_arquivo,
        read_only=False,
        data_only=True
    )

    sheet_periodo = None

    for nome in wb.sheetnames:
        if SHEET_PERIODO_REGEX.search(nome):
            sheet_periodo = wb[nome]
            break

    if sheet_periodo is None:
        wb.close()
        raise Exception("Sheet de período não encontrada")

    # Remove TODAS as outras sheets
    for nome in wb.sheetnames[:]:
        if wb[nome] != sheet_periodo:
            wb.remove(wb[nome])

    # Renomeia
    sheet_periodo.title = "Plan1"

    # Remove filtros
    sheet_periodo.auto_filter = None

    wb.save(caminho_arquivo)
    wb.close()


def processar_arquivos():
    print("\n========================================")
    print(" PROCESSANDO FALHAS OPERACIONAIS ITAU ")
    print("========================================\n")

    arquivos = [
        f for f in DOWNLOADS.iterdir()
        if f.is_file() and ARQUIVO_REGEX.match(f.name)
    ]

    if not arquivos:
        print("❌ Nenhum arquivo encontrado.")
        return

    print(f"✔ {len(arquivos)} arquivo(s) encontrado(s):\n")

    for arquivo in arquivos:
        print(f"➡ Processando: {arquivo.name}")

        try:
            # Trata LOCALMENTE (muito mais rápido)
            tratar_excel(arquivo)
            print("   ✔ Sheet correta mantida e renomeada")

            # Move para a rede depois de tratado
            destino_final = DESTINO_REDE / arquivo.name
            shutil.move(str(arquivo), str(destino_final))
            print("   ✔ Movido para a rede")

        except Exception as e:
            print(f"   ❌ Erro: {e}")

    print("\n✔ Finalizado.\n")


# -------------------------------
# Execução
# -------------------------------
if __name__ == "__main__":
    processar_arquivos()
