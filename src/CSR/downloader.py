import os
from openpyxl import load_workbook


def ajustar_nome_sheet(caminho_arquivo):
    """
    Renomeia a sheet 'Plan1' para 'Planilha1'.
    Se 'Planilha1' já existir, não faz nada.
    """

    try:
        wb = load_workbook(caminho_arquivo)

        sheets = wb.sheetnames

        if "Planilha1" in sheets:
            return  # já está correto

        if "Plan1" in sheets:
            wb["Plan1"].title = "Planilha1"
            wb.save(caminho_arquivo)
            print(f"📝 Sheet ajustada para Planilha1: {caminho_arquivo}")

    except Exception as e:
        print(f"⚠️ Erro ao ajustar sheet em {caminho_arquivo}: {e}")


def salvar_anexos(emails, output_path):
    """
    Salva somente anexos .xlsx dos e-mails recebidos.
    Após salvar, ajusta o nome da sheet.
    """

    os.makedirs(output_path, exist_ok=True)

    for msg in emails:
        if msg.Attachments.Count == 0:
            print(f"⚠️ Email sem anexos: {msg.Subject}")
            continue

        for attachment in msg.Attachments:
            nome_arquivo = attachment.FileName

            if not nome_arquivo or "." not in nome_arquivo:
                print(f"⏭️ Ignorado (sem nome): {nome_arquivo}")
                continue

            if not nome_arquivo.lower().endswith(".xlsx"):
                print(f"⏭️ Ignorado (não é Excel): {nome_arquivo}")
                continue

            destino = os.path.join(output_path, nome_arquivo)

            try:
                attachment.SaveAsFile(destino)
                print(f"📁 Arquivo salvo: {destino}")

                # 🔥 Ajusta o nome da sheet
                ajustar_nome_sheet(destino)

            except Exception as e:
                print(f"❌ Erro ao salvar {nome_arquivo}: {e}")
